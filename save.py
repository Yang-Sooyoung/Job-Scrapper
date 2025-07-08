from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def save_to_file(jobs, source="jobs"):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{source.capitalize()} Jobs"

    headers = ["title", "company", "location", "deadline"]
    ws.append(headers)

    for job in jobs:
        title = job.get("title", "")
        company = job.get("company", "")
        location = job.get("location", "")
        deadline = job.get("deadline", "")
        link = job.get("link", "")

        cell = ws.cell(row=ws.max_row + 1, column=1, value=title)
        if link:
            cell.hyperlink = link
            cell.font = Font(color="0000FF", underline="single")

        ws.cell(row=cell.row, column=2, value=company)
        ws.cell(row=cell.row, column=3, value=location)
        ws.cell(row=cell.row, column=4, value=deadline)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    filename = f"jobs_{source}.xlsx"
    wb.save(filename)
    print(f"Excel 저장 완료: {filename}")